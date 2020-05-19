from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import load_workbook
import xlrd
from datetime import date

location_list = []
numbers_list = []
numberTotalJobs = 0
cl = [] ##companyList
jl = [] ##JobList
ll = [] ##LocationList

##appends the scraped data to a spreadsheet
#Job/Company/Location
def appendToCompanyJobs_Workbook(listToAppend_JL, listToAppend_CL, listToAppend_LL):
    path = r"C:\Users\...\Softwareentwickler\CompanyJobs.xlsx" #Please create a spreadsheet called "CompanyJobs.xlsx" and insert the location here!
    book = load_workbook(path)
    sheet_name = str(date.today())
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book

    CompanyJob = pd.DataFrame({'Unternehmen' : listToAppend_CL, 'Job' : listToAppend_JL, 'Location' : listToAppend_LL})
    CompanyJob.to_excel(writer, sheet_name=sheet_name, index=True)
    writer.save()
    writer.close()

#How many Jobs are in which region
def appendToLocationCount_Workbook():
    path = r"C:\Users\...\Softwareentwickler\LocationCount.xlsx" #Please create a spreadsheet called "LocationCount.xlsx" and insert the location here!
    book = load_workbook(path)
    sheet_name = str(date.today())
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book

    numberJobs = pd.DataFrame({'Ort' : location_list, 'Anzahl': numbers_list})
    numberJobs.to_excel(writer, sheet_name=sheet_name, index=True)
    writer.save()
    writer.close()

#How many total Jobs on which date
def append_df_to_excel(df, sheet_name='Sheet1', startrow=1, truncate_sheet=False, **to_excel_kwargs):
    path = r"C:\Users\...\Softwareentwickler\NumberOfJobs_Date.xlsx"

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(path, engine='openpyxl')
    try:
        # try to open an existing workbook
        writer.book = load_workbook(path)
        startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
            pass
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, header=False, index=True)
    writer.save()

##Scrapers
#location-number Scraper
def locationNumbers(data):
    soup = BeautifulSoup(data.text, 'html.parser')
    rbLocation = soup.find_all('span', attrs={'class': "rbLabel"})
    rbNumber = soup.find_all('span', attrs={'class': "rbCount"})
    for l in rbLocation:
        locations = l.text.strip()
        location_list.append(locations)

    for n in rbNumber:
        numbers = n.text.strip()
        numbers_list.append(numbers)

    if len(rbNumber) == 0:
        rbNumber = soup.find_all('span', attrs={'class': "rbCount"})
        for n in rbNumber:
            numbers = n.text.strip()
            numbers_list.append(numbers)

    print('LocationList: ' + str(len(location_list)))
    print('NumbersList: ' + str(len(numbers_list)))
    if len(location_list) == len(numbers_list):
        appendToLocationCount_Workbook()
    else:
        print('not the same content: LocationList, NumbersList')

#Number of total Jobs Scrapers
def getTotalNumber(data):
    dateToday = str(date.today())
    soup = BeautifulSoup(data.text, 'html.parser')
    totalNumber = soup.find('meta', attrs={'name':"description"})
    totalNumberStr = str(totalNumber).split()
    res = [int(i) for i in totalNumberStr if i.isdigit()]
    finalTotalNumber = res[:-1]
    for Lumber in finalTotalNumber:
        numberTotalJobs = Lumber
    dataframe = pd.DataFrame({'Date' : dateToday, 'No Of Jobs - SoftwareEntwickler': finalTotalNumber})
    #append_df_to_excel(df=dataframe, sheet_name='Sheet1')

    return numberTotalJobs

#Company Scraper
def companyScrape(data):
    soup = BeautifulSoup(data.text, 'html.parser')
    company = soup.find_all('span', attrs={'class':"company"})
    company_list = []
    for c in company:
        companyStr = c.text.strip()
        company_list.append(companyStr)

    return company_list

#Job Scraper
def jobScrape(data):
    soup = BeautifulSoup(data.text, 'html.parser')
    job = soup.find_all('a', attrs={'data-tn-element':"jobTitle"})
    job_list = []
    for j in job:
        jobStr = j.text.strip()
        job_list.append(jobStr)

    return job_list


#Location Scraper
def locationScrape(data):
    soup = BeautifulSoup(data.text, 'html.parser')
    location = soup.find_all('span', attrs={'class':"location accessible-contrast-color-location"})
    location_list = []
    for l in location:
        locationStr = l.text.strip()
        location_list.append(locationStr)

    return location_list

##Executer Function
def jobsOut_Func():
    job = input('What Job are you looking for: ')
    location = input('Where are you looking for it: ')
    url = 'https://de.indeed.com/jobs?q=' + job + '&l=' + location
    counter = 0

    #Scrapes the first page
    if counter == 0:
        data = requests.get(url)
        numberTotalJobs = getTotalNumber(data)
        print('Total Number of Jobs = ' + str(numberTotalJobs))
        locationNumbers(data)
        cl.extend(companyScrape(data))
        jl.extend(jobScrape(data))
        ll.extend(locationScrape(data))

    #Loop for the following pages
    if numberTotalJobs > 0:
        while counter <= numberTotalJobs:
            counter += 10
            url_2 = url + str('&start=') + str(counter)
            data = requests.get(url_2)
            cl.extend(companyScrape(data))
            jl.extend(jobScrape(data))
            ll.extend(locationScrape(data))
    counter +=10

    diffNumberTotaJobs = len(cl) - numberTotalJobs
    listToAppend_JL = jl[:-diffNumberTotaJobs]
    listToAppend_CL = cl[:-diffNumberTotaJobs]
    listToAppend_LL = ll[:-diffNumberTotaJobs]

    #Appends the Scrapes to the workbook
    if len(listToAppend_CL) == len(listToAppend_JL):
        appendToCompanyJobs_Workbook(listToAppend_JL, listToAppend_CL, listToAppend_LL)

    print("ComapnyList: " + str(len(listToAppend_CL)))
    print('Joblist: ' + str(len(listToAppend_JL)))


jobsOut_Func()
